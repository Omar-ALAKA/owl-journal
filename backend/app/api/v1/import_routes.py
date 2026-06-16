# app/api/v1/import_routes.py
import csv
import io
import logging
import re
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, Request, UploadFile, File, Form
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from app.database import get_db
from app.models.trade import Trade
from app.models.account import Account

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/import", tags=["import"])

# ───────────────────────────────────────────────────────────
# Multilingual section markers (MT5 exports have 3 sections)
# ───────────────────────────────────────────────────────────
SECTION_MARKERS = {
    "positions": [
        "positions", "posiciones", "posições", "positionen", "positionen",
        "pozice", "pozycje", "Διαθέσιμες θέσεις", "مواضع",
    ],
    "orders": [
        "orders", "ordres", "órdenes", "ordens", "aufträge",
        "commandes", "comandos", "pedidos", "orδερζ", "aufträge",
        # Pending orders variants
        "pending orders", "ordres en attente", "órdenes pendientes",
        "ordens pendentes", "ausstehende aufträge",
    ],
    "deals": [
        "deals", "transactions", "operaciones", "negócios",
        "transaktionen", "opérations", "transacties", "operacje",
        "συναλλαγές", "معاملات", "transaction history",
    ],
}


def _normalize_header(header: str) -> str:
    if header is None:
        return ""
    return re.sub(r"\s+", " ", header.strip().lower())


def _is_section_marker(row) -> str | None:
    """Return section type if the row is a section marker, else None."""
    # A marker row typically has only 1 non-null cell (the title)
    non_null = [c for c in row if c is not None and str(c).strip()]
    if len(non_null) != 1:
        return None
    label = _normalize_header(str(non_null[0]))
    for section_type, variants in SECTION_MARKERS.items():
        if label in variants:
            return section_type
    return None


def _map_columns(headers: list[str]) -> dict[str, int]:
    normalized = [_normalize_header(h) for h in headers]

    COLUMN_MAP = {
        "ticket": ["ticket", "order", "order id", "trade id", "deal",
                   "#", "trade #", "deal id", "position"],
        "open_time": [
            "open time", "opentime", "open_time", "time", "date/time",
            "opening time", "open", "datetime", "timestamp", "time opened",
            "open date", "date opened", "heure", "date d'ouverture",
            "heure d'ouverture",
        ],
        "close_time": [
            "close time", "closetime", "close_time", "closing time",
            "close", "time closed", "closing date", "date closed",
            "heure de fermeture", "date de fermeture",
            "heure",
        ],
        "symbol": ["symbol", "instrument", "pair", "ticker", "market", "asset", "symbole"],
        "direction": [
            "direction", "type", "side", "trade type", "order type",
            "transaction type", "cmd", "operation", "trade direction",
        ],
        "volume": [
            "volume", "size", "lots", "quantity", "qty", "amount",
            "volume (lots)", "size (lots)", "volume(lots)", "lot",
        ],
        "entry_price": [
            "entry price", "entry_price", "open price", "price",
            "opening price", "entry", "open px", "buy price", "sell price",
            "prix", "prix d'entrée", "prix ouverture",
        ],
        "exit_price": [
            "exit price", "exit_price", "close price", "closing price",
            "exit", "close price", "close px", "prix de sortie",
            "prix fermeture", "prix",
        ],
        "sl_price": [
            "sl", "stop loss", "sl_price", "stoploss", "stop loss price",
            "sl price", "stop loss (sl)", "s / l",
        ],
        "tp_price": [
            "tp", "take profit", "tp_price", "takeprofit", "take profit price",
            "tp price", "take profit (tp)", "t / p",
        ],
        "commission": ["commission", "comm", "commissions", "fee", "fees"],
        "swap": ["swap", "swap charge", "rollover", "echange"],
        "profit": [
            "profit", "pnl", "p&l", "net pnl", "gain", "pnl (usd)",
            "profit (usd)", "net profit", "result", "pnl (result)",
            "realized p&l", "closed p/l",
        ],
        "r_multiple": ["r multiple", "r_multiple", "r:r", "r/r", "rr"],
        "session": ["session", "trading session", "session name"],
        "setup": ["setup", "strategy", "setup name", "trade setup"],
        "notes": ["notes", "comment", "comments", "remark", "remarks"],
    }

    mapping: dict[str, int] = {}
    used_indices: set[int] = set()

    for canonical, variants in COLUMN_MAP.items():
        for idx, norm in enumerate(normalized):
            if norm in variants and canonical not in mapping and idx not in used_indices:
                mapping[canonical] = idx
                used_indices.add(idx)
                break

    # Second pass: handle duplicate headers (e.g. "Heure" x2 in Equity Edge)
    for canonical, variants in COLUMN_MAP.items():
        if canonical not in mapping:
            for idx, norm in enumerate(normalized):
                if norm in variants and idx not in used_indices:
                    mapping[canonical] = idx
                    used_indices.add(idx)
                    break

    return mapping


def _safe_float(val: Any) -> float:
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s or s in ("-", "—", "N/A", "n/a"):
        return 0.0
    s = re.sub(r"[£€$¥₹,\s]", "", s)
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    try:
        return float(s)
    except (ValueError, TypeError):
        return 0.0


def _safe_datetime(val: Any, fmt_hints: list[str] | None = None) -> datetime | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    # Excel serial date
    if isinstance(val, float):
        try:
            if 1 < val < 200000:
                return datetime(1899, 12, 30) + timedelta(days=int(val))
        except (ValueError, OverflowError):
            pass
    s = str(val).strip()
    if not s or s in ("-", "—", "N/A", "n/a"):
        return None
    formats = fmt_hints or [
        "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M",
        "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%dT%H:%M",
        "%Y-%m-%d",
        "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y",
        "%m/%d/%Y %H:%M:%S", "%m/%d/%Y %H:%M", "%m/%d/%Y",
        "%d.%m.%Y %H:%M:%S", "%d.%m.%Y %H:%M", "%d.%m.%Y",
        "%Y.%m.%d %H:%M:%S", "%Y.%m.%d",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _detect_direction(val: Any) -> str:
    if val is None:
        return ""
    s = str(val).strip().lower()
    if s in ("buy", "long", "b", "1", "buy market"):
        return "long"
    if s in ("sell", "short", "s", "-1", "2", "sell market"):
        return "short"
    return s


def _detect_format(headers: list[str]) -> str:
    normalized = [_normalize_header(h) for h in headers]
    joined = " ".join(normalized)

    # Equity Edge (French): Symbole + S/L + T/P + Echange
    if "symbole" in joined and ("s / l" in joined or "t / p" in joined) and "echange" in joined:
        return "equity_edge"
    if "signal" in joined or "equity edge" in joined:
        return "equity_edge"

    # MT5: Deal + Symbol + Type + Volume + Price + SL + TP + Profit + Commission + Swap
    if any(h in normalized for h in ["deal", "deal id"]) and "profit" in joined:
        if "swap" in joined or "commission" in joined:
            return "mt5"

    # FTMO
    if "deal" in joined and "swap" in joined and "commission" in joined:
        return "ftmo"

    # MyFundedFX
    if "ticket" in joined and ("open time" in joined or "opening time" in joined):
        return "myfundedfx"

    # FundedNext
    if "order" in joined and "profit" in joined and "swap" in joined:
        return "fundednext"

    return "generic"


def _cell_to_str(cell: Any) -> str:
    """Convert an openpyxl cell value to string."""
    if cell is None:
        return ""
    if isinstance(cell, datetime):
        return cell.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(cell, float):
        if cell == int(cell):
            return str(int(cell))
        return str(cell)
    return str(cell).strip()


def _parse_csv_content(content: str, tz_code: str = "UTC", session_config: dict | None = None) -> list[dict]:
    """Parse CSV content and return list of trade dicts."""
    reader = csv.reader(io.StringIO(content))
    rows = list(reader)
    if len(rows) < 2:
        return []

    headers = rows[0]
    col_map = _map_columns(headers)
    fmt = _detect_format(headers)

    trades = []
    for row in rows[1:]:
        if not row or all(not c.strip() for c in row):
            continue

        def get(col_name: str, default: str = "") -> str:
            idx = col_map.get(col_name)
            if idx is None or idx >= len(row):
                return default
            return row[idx].strip()

        open_time_raw = get("open_time")
        close_time_raw = get("close_time")
        open_time = _safe_datetime(open_time_raw)
        close_time = _safe_datetime(close_time_raw)

        if not open_time:
            continue

        symbol = get("symbol")
        if not symbol:
            continue

        direction = _detect_direction(get("direction"))
        volume = _safe_float(get("volume"))
        entry_price = _safe_float(get("entry_price"))
        exit_price = _safe_float(get("exit_price"))
        sl_price = _safe_float(get("sl_price"))
        tp_price = _safe_float(get("tp_price"))
        commission = _safe_float(get("commission"))
        swap = _safe_float(get("swap"))
        profit = _safe_float(get("profit"))
        r_multiple = _safe_float(get("r_multiple"))
        session = get("session")
        setup = get("setup")
        notes = get("notes")
        ticket = get("ticket")

        # Auto-detect session from open_time using timezone + user session config
        if not session and open_time:
            try:
                from app.services.timezone import get_offset
                from app.services.session_config import get_session_for_time
                offset = get_offset(tz_code, open_time)
                session = get_session_for_time(open_time, offset, session_config)
            except Exception:
                # Fallback: detect session from UTC hour
                h = open_time.hour
                if 0 <= h < 8:
                    session = "Asia"
                elif 8 <= h < 13:
                    session = "London"
                elif 13 <= h < 22:
                    session = "New York"
                else:
                    session = "Off-hours"

        if not direction and profit != 0:
            direction = "long" if profit > 0 else "short"

        trade = {
            "ticket": ticket or None,
            "open_time": open_time.isoformat(),
            "close_time": close_time.isoformat() if close_time else None,
            "symbol": symbol.upper(),
            "direction": direction or "long",
            "volume": volume,
            "entry_price": entry_price,
            "exit_price": exit_price if exit_price > 0 else None,
            "sl_price": sl_price if sl_price > 0 else None,
            "tp_price": tp_price if tp_price > 0 else None,
            "commission": commission,
            "swap": swap,
            "profit": profit,
            "r_multiple": r_multiple if r_multiple > 0 else None,
            "session": session or None,
            "setup": setup or None,
            "notes": notes or None,
            "detected_format": fmt,
        }
        trades.append(trade)

    return trades


def _parse_xlsx_content(content_bytes: bytes, tz_code: str = "UTC", session_config: dict | None = None) -> list[dict]:
    """Parse XLSX content directly using openpyxl.

    Handles multi-section MT5 exports (Positions / Orders / Deals)
    by detecting section markers in multiple languages and only
    parsing the Positions block.
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content_bytes), read_only=True, data_only=True)
    all_trades = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 2:
            continue

        # ── Step 1: Detect section boundaries ──
        # MT5 exports have 3 sections: Positions, Orders, Deals
        # We only want the Positions block.
        positions_header_row = None
        positions_start = None
        positions_end = None

        for i, row in enumerate(rows):
            marker = _is_section_marker(row)
            if marker == "positions":
                positions_header_row = i + 1  # next row = headers
                positions_start = i + 2       # data starts after headers
            elif marker in ("orders", "deals") and positions_start is not None:
                positions_end = i
                break

        # ── Step 2: Slice to Positions block only ──
        if positions_start and positions_end and positions_header_row is not None:
            header_row_raw = rows[positions_header_row]
            header_row = [_cell_to_str(h) for h in header_row_raw]
            data_rows = rows[positions_start:positions_end]
            logger.info(
                f"XLSX sheet '{sheet_name}': Positions section detected "
                f"(rows {positions_start}-{positions_end}, {len(data_rows)} data rows)"
            )
        else:
            # Fallback: auto-detect header row
            header_row_idx = 0
            for i, row in enumerate(rows):
                non_empty = sum(1 for c in row if c is not None and str(c).strip())
                if non_empty >= 3:
                    header_row_idx = i
                    break
            header_row_raw = rows[header_row_idx]
            header_row = [_cell_to_str(h) for h in header_row_raw]
            data_rows = rows[header_row_idx + 1:]
            logger.info(
                f"XLSX sheet '{sheet_name}': no section markers found, "
                f"using fallback header at row {header_row_idx}"
            )

        logger.info(f"XLSX headers: {header_row}")
        fmt = _detect_format(header_row)
        col_map = _map_columns(header_row)
        logger.info(f"XLSX column mapping ({len(col_map)} cols): {col_map}")

        # ── Step 3: Parse data rows ──
        trades_parsed = 0
        trades_skipped = 0

        for row in data_rows:
            if all(c is None for c in row):
                continue

            str_row = [_cell_to_str(c) for c in row]

            def get(col_name: str, default: str = "") -> str:
                idx = col_map.get(col_name)
                if idx is None or idx >= len(str_row):
                    return default
                return str_row[idx]

            open_time_raw = get("open_time")
            close_time_raw = get("close_time")
            open_time = _safe_datetime(open_time_raw)
            close_time = _safe_datetime(close_time_raw)

            if not open_time:
                trades_skipped += 1
                continue

            symbol = get("symbol")
            if not symbol:
                trades_skipped += 1
                continue

            direction = _detect_direction(get("direction"))
            volume = _safe_float(get("volume"))
            entry_price = _safe_float(get("entry_price"))
            exit_price = _safe_float(get("exit_price"))
            sl_price = _safe_float(get("sl_price"))
            tp_price = _safe_float(get("tp_price"))
            commission = _safe_float(get("commission"))
            swap = _safe_float(get("swap"))
            profit = _safe_float(get("profit"))
            r_multiple = _safe_float(get("r_multiple"))
            session = get("session")
            setup = get("setup")
            notes = get("notes")
            ticket = get("ticket")

            # Auto-detect session from open_time using timezone + user session config
            if not session and open_time:
                try:
                    from app.services.timezone import get_offset
                    from app.services.session_config import get_session_for_time
                    offset = get_offset(tz_code, open_time)
                    session = get_session_for_time(open_time, offset, session_config)
                except Exception:
                    # Fallback: detect session from UTC hour
                    h = open_time.hour
                    if 0 <= h < 8:
                        session = "Asia"
                    elif 8 <= h < 13:
                        session = "London"
                    elif 13 <= h < 22:
                        session = "New York"
                    else:
                        session = "Off-hours"

            if not direction and profit != 0:
                direction = "long" if profit > 0 else "short"

            trade = {
                "ticket": ticket or None,
                "open_time": open_time.isoformat(),
                "close_time": close_time.isoformat() if close_time else None,
                "symbol": symbol.upper(),
                "direction": direction or "long",
                "volume": volume,
                "entry_price": entry_price,
                "exit_price": exit_price if exit_price > 0 else None,
                "sl_price": sl_price if sl_price > 0 else None,
                "tp_price": tp_price if tp_price > 0 else None,
                "commission": commission,
                "swap": swap,
                "profit": profit,
                "r_multiple": r_multiple if r_multiple > 0 else None,
                "session": session or None,
                "setup": setup or None,
                "notes": notes or None,
                "detected_format": fmt,
            }
            all_trades.append(trade)
            trades_parsed += 1

        logger.info(
            f"XLSX sheet '{sheet_name}': parsed {trades_parsed} trades, "
            f"skipped {trades_skipped} rows"
        )

    logger.info(f"XLSX total: {len(all_trades)} trades from {len(wb.sheetnames)} sheet(s)")
    return all_trades


# ───────────────────────────────────────────────────────────
# POST /import/preview
# ───────────────────────────────────────────────────────────
@router.post("/preview")
async def preview_import(
    request: Request,
    file: UploadFile = File(...),
    account_id: int = Form(None),
    db: AsyncSession = Depends(get_db),
):
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")

    logger.info(f"Import preview: file={file.filename}")
    filename = file.filename.lower()

    # Get timezone from request header
    tz_code = request.headers.get("X-Timezone", "UTC")

    content_bytes = await file.read()
    logger.info(f"Import preview: file={file.filename}, size={len(content_bytes)}")
    if len(content_bytes) > 10 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="File too large (max 10MB)")

    # Get session config from account if available
    session_config = None
    account_info = None
    if account_id:
        acc_result = await db.execute(
            select(Account).where(Account.id == account_id)
        )
        acc = acc_result.scalar_one_or_none()
        if acc:
            session_config = acc.session_hours
            account_info = {
                "id": acc.id,
                "name": acc.name,
                "broker": acc.broker,
            }

    if filename.endswith(".csv") or filename.endswith(".txt"):
        for encoding in ("utf-8", "utf-8-sig", "latin-1", "cp1252"):
            try:
                content = content_bytes.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            raise HTTPException(400, detail="Unable to decode file")
        trades = _parse_csv_content(content, tz_code=tz_code, session_config=session_config)

    elif filename.endswith((".xlsx", ".xls")):
        try:
            trades = _parse_xlsx_content(content_bytes, tz_code=tz_code, session_config=session_config)
        except ImportError:
            raise HTTPException(
                500,
                detail="openpyxl not installed. Install with: pip install openpyxl",
            )
    else:
        raise HTTPException(
            400,
            detail="Unsupported file format. Use .csv, .txt, .xlsx, or .xls",
        )

    account_info = account_info or None

    total = len(trades)
    detected_format = trades[0]["detected_format"] if trades else "unknown"
    symbols = list(set(t["symbol"] for t in trades if t["symbol"]))
    total_profit = sum(t["profit"] for t in trades)
    wins = sum(1 for t in trades if t["profit"] > 0)
    losses = sum(1 for t in trades if t["profit"] < 0)

    logger.info(
        f"Import preview result: format={detected_format}, trades={total}, "
        f"profit={total_profit:.2f}, wins={wins}, losses={losses}"
    )

    return {
        "preview": True,
        "filename": file.filename,
        "detected_format": detected_format,
        "total_trades": total,
        "symbols": sorted(symbols),
        "stats": {
            "total_profit": round(total_profit, 2),
            "wins": wins,
            "losses": losses,
            "win_rate": round(wins / total * 100, 1) if total > 0 else 0,
        },
        "account": account_info,
        "trades": trades,
    }


# ───────────────────────────────────────────────────────────
# POST /import/confirm
# ───────────────────────────────────────────────────────────
@router.post("/confirm")
async def confirm_import(
    data: dict,
    db: AsyncSession = Depends(get_db),
):
    trades_data = data.get("trades", [])
    account_id = data.get("account_id")

    logger.info(f"Import confirm: account_id={account_id}, trades={len(trades_data)}")

    if not trades_data:
        raise HTTPException(status_code=400, detail="No trades to import")

    if not account_id:
        raise HTTPException(status_code=400, detail="account_id is required")

    acc_result = await db.execute(
        select(Account).where(Account.id == account_id)
    )
    account = acc_result.scalar_one_or_none()
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")

    inserted = 0
    errors = []
    created_trades = []

    for idx, t in enumerate(trades_data):
        try:
            symbol = t.get("symbol")
            open_time_str = t.get("open_time")
            direction = t.get("direction", "long")

            if not symbol or not open_time_str:
                errors.append({
                    "row": idx,
                    "error": "Missing required fields: symbol, open_time",
                })
                continue

            if isinstance(open_time_str, str):
                open_time = _safe_datetime(open_time_str)
            elif isinstance(open_time_str, datetime):
                open_time = open_time_str
            else:
                open_time = None

            if not open_time:
                errors.append({
                    "row": idx,
                    "error": f"Cannot parse open_time: {open_time_str}",
                })
                continue

            close_time = None
            close_time_str = t.get("close_time")
            if close_time_str:
                if isinstance(close_time_str, str):
                    close_time = _safe_datetime(close_time_str)
                elif isinstance(close_time_str, datetime):
                    close_time = close_time_str

            profit = _safe_float(t.get("profit", 0))
            is_winner = 1 if profit > 0 else 0

            r_multiple = t.get("r_multiple")
            if r_multiple is None and profit != 0:
                sl_price = _safe_float(t.get("sl_price", 0))
                entry_price = _safe_float(t.get("entry_price", 0))
                if sl_price > 0 and entry_price > 0:
                    risk = abs(entry_price - sl_price)
                    if risk > 0:
                        r_multiple = profit / risk

            trade = Trade(
                account_id=account_id,
                ticket=t.get("ticket"),
                open_time=open_time,
                close_time=close_time,
                symbol=symbol.upper(),
                direction=direction,
                volume=_safe_float(t.get("volume", 0)),
                entry_price=_safe_float(t.get("entry_price", 0)),
                exit_price=_safe_float(t.get("exit_price", 0)) or None,
                sl_price=_safe_float(t.get("sl_price", 0)) or None,
                tp_price=_safe_float(t.get("tp_price", 0)) or None,
                commission=_safe_float(t.get("commission", 0)),
                swap=_safe_float(t.get("swap", 0)),
                profit=profit,
                r_multiple=r_multiple if r_multiple else 0,
                session=t.get("session"),
                setup=t.get("setup"),
                notes=t.get("notes"),
                is_winner=is_winner,
            )
            db.add(trade)
            await db.flush()
            created_trades.append(trade.id)
            inserted += 1

        except Exception as e:
            logger.error(f"Import error row {idx}: {e}", exc_info=True)
            errors.append({"row": idx, "error": str(e)})

    logger.info(f"Import confirm done: inserted={inserted}, errors={len(errors)}")

    return {
        "message": f"Imported {inserted} trades",
        "account_id": account_id,
        "inserted": inserted,
        "errors": errors,
        "error_count": len(errors),
        "trade_ids": created_trades,
    }
